import os, sqlite3, uuid, json
from functools import wraps
from io import BytesIO
from datetime import date

from flask import (Flask, render_template, request, redirect, url_for,
                   flash, g, session, send_file, jsonify)
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
DB_PATH   = os.path.join(BASE_DIR, 'fichas.db')
UPLOADS   = os.path.join(BASE_DIR, 'static', 'uploads')
EXT_IMG   = {'png','jpg','jpeg','gif','webp'}
EXT_DOC   = {'pdf','jpg','jpeg','png','doc','docx','xls','xlsx'}

ADMIN_USER     = os.environ.get('ADMIN_USER', 'admin')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'admin123')

app = Flask(__name__)
app.config['SECRET_KEY']        = os.environ.get('SECRET_KEY', 'cambia-esta-clave')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

os.makedirs(UPLOADS, exist_ok=True)

# ─────────────────────────────────────────────
#  CONFIGURACIÓN DE SECCIONES
# ─────────────────────────────────────────────
# Cada sección define:
#   nombre       → título visible
#   columnas     → [(header, clave)] donde clave es campo base o clave de datos_extra
#   campos       → [{name, label, tipo, ancho, opciones?}]  para el modal
# Campos base en la tabla: tipo_documento, numero_documento, fecha_documento,
#                          servicio_documento, estado_tramite
SECCIONES_CONFIG = {
    'designaciones': {
        'nombre': 'Designaciones',
        'columnas': [
            ('Tipo Documento',   'tipo_documento'),
            ('N° Documento',     'numero_documento'),
            ('Fecha Documento',  'fecha_documento'),
            ('Servicio',         'servicio_documento'),
            ('Fecha Desde',      'fecha_desde'),
            ('Fecha Hasta',      'fecha_hasta'),
            ('Calidad',          'calidad'),
            ('Horas',            'horas'),
            ('Estado Trámite',   'estado_tramite'),
        ],
        'campos': [
            {'name':'tipo_documento',   'label':'Tipo Documento',   'tipo':'text',   'base':True},
            {'name':'numero_documento', 'label':'N° Documento',     'tipo':'text',   'base':True},
            {'name':'fecha_documento',  'label':'Fecha Documento',  'tipo':'date',   'base':True},
            {'name':'servicio_documento','label':'Servicio Desempeño','tipo':'text', 'base':True},
            {'name':'fecha_desde',      'label':'Fecha Desde',      'tipo':'date'},
            {'name':'fecha_hasta',      'label':'Fecha Hasta',      'tipo':'date'},
            {'name':'calidad',          'label':'Calidad',          'tipo':'text'},
            {'name':'horas',            'label':'Horas',            'tipo':'number'},
            {'name':'estado_tramite',   'label':'Estado Trámite',   'tipo':'text',   'base':True},
        ],
    },
    'cese_funciones': {
        'nombre': 'Cese de Funciones',
        'columnas': [
            ('Tipo Documento',  'tipo_documento'),
            ('N° Documento',    'numero_documento'),
            ('Fecha Documento', 'fecha_documento'),
            ('Servicio',        'servicio_documento'),
            ('Tipo Cese',       'tipo_cese'),
            ('Causal del Cese', 'causal_cese'),
            ('Horas Cesadas',   'horas_cesadas'),
            ('Fecha Desde',     'fecha_desde'),
            ('Estado Trámite',  'estado_tramite'),
        ],
        'campos': [
            {'name':'tipo_documento',   'label':'Tipo Documento',  'tipo':'text',  'base':True},
            {'name':'numero_documento', 'label':'N° Documento',    'tipo':'text',  'base':True},
            {'name':'fecha_documento',  'label':'Fecha Documento', 'tipo':'date',  'base':True},
            {'name':'servicio_documento','label':'Servicio',       'tipo':'text',  'base':True},
            {'name':'tipo_cese',        'label':'Tipo Cese',       'tipo':'text'},
            {'name':'causal_cese',      'label':'Causal del Cese', 'tipo':'text'},
            {'name':'horas_cesadas',    'label':'Horas Cesadas',   'tipo':'number'},
            {'name':'fecha_desde',      'label':'Fecha Desde',     'tipo':'date'},
            {'name':'estado_tramite',   'label':'Estado Trámite',  'tipo':'text',  'base':True},
        ],
    },
    'responsabilidad_administrativa': {
        'nombre': 'Responsabilidad Administrativa',
        'columnas': [
            ('Tipo Documento',  'tipo_documento'),
            ('N° Documento',    'numero_documento'),
            ('Fecha Documento', 'fecha_documento'),
            ('Servicio',        'servicio_documento'),
            ('Materia',         'materia'),
            ('Conclusión',      'conclusion'),
            ('Estado Trámite',  'estado_tramite'),
        ],
        'campos': [
            {'name':'tipo_documento',   'label':'Tipo Documento',  'tipo':'text', 'base':True},
            {'name':'numero_documento', 'label':'N° Documento',    'tipo':'text', 'base':True},
            {'name':'fecha_documento',  'label':'Fecha Documento', 'tipo':'date', 'base':True},
            {'name':'servicio_documento','label':'Servicio',       'tipo':'text', 'base':True},
            {'name':'materia',          'label':'Materia',         'tipo':'text'},
            {'name':'conclusion',       'label':'Conclusión',      'tipo':'text'},
            {'name':'estado_tramite',   'label':'Estado Trámite',  'tipo':'text', 'base':True},
        ],
    },
    'inhabilidades': {
        'nombre': 'Inhabilidades',
        'columnas': [
            ('Tipo Documento',     'tipo_documento'),
            ('N° Documento',       'numero_documento'),
            ('Fecha Documento',    'fecha_documento'),
            ('Servicio',           'servicio_documento'),
            ('Materia',            'materia'),
            ('Fecha Desde',        'fecha_desde'),
            ('Tiempo Inhabilidad', 'tiempo_inhabilidad'),
            ('Estado Trámite',     'estado_tramite'),
        ],
        'campos': [
            {'name':'tipo_documento',    'label':'Tipo Documento',     'tipo':'text', 'base':True},
            {'name':'numero_documento',  'label':'N° Documento',       'tipo':'text', 'base':True},
            {'name':'fecha_documento',   'label':'Fecha Documento',    'tipo':'date', 'base':True},
            {'name':'servicio_documento','label':'Servicio',           'tipo':'text', 'base':True},
            {'name':'materia',           'label':'Materia',            'tipo':'text'},
            {'name':'fecha_desde',       'label':'Fecha Desde',        'tipo':'date'},
            {'name':'tiempo_inhabilidad','label':'Tiempo Inhabilidad', 'tipo':'text'},
            {'name':'estado_tramite',    'label':'Estado Trámite',     'tipo':'text', 'base':True},
        ],
    },
    'estudios': {
        'nombre': 'Estudios',
        'columnas': [
            ('Tipo Documento',    'tipo_documento'),
            ('N° Documento',      'numero_documento'),
            ('Fecha Documento',   'fecha_documento'),
            ('Nivel de Estudio',  'nivel_estudio'),
            ('Título / Carrera',  'titulo_carrera'),
            ('Fecha Otorgamiento','fecha_otorgamiento'),
            ('Institución',       'institucion'),
            ('Estado Trámite',    'estado_tramite'),
        ],
        'campos': [
            {'name':'tipo_documento',    'label':'Tipo Documento',     'tipo':'text', 'base':True},
            {'name':'numero_documento',  'label':'N° Documento',       'tipo':'text', 'base':True},
            {'name':'fecha_documento',   'label':'Fecha Documento',    'tipo':'date', 'base':True},
            {'name':'servicio_documento','label':'Servicio',           'tipo':'text', 'base':True},
            {'name':'nivel_estudio',     'label':'Nivel de Estudio',   'tipo':'text'},
            {'name':'titulo_carrera',    'label':'Título / Carrera',   'tipo':'text'},
            {'name':'fecha_otorgamiento','label':'Fecha Otorgamiento', 'tipo':'date'},
            {'name':'fecha_convalidacion','label':'Fecha Convalidación','tipo':'date'},
            {'name':'institucion',       'label':'Institución',        'tipo':'text'},
            {'name':'estado_tramite',    'label':'Estado Trámite',     'tipo':'text', 'base':True},
        ],
    },
    'contrato_honorarios': {
        'nombre': 'Contrato a Honorarios',
        'columnas': [
            ('Tipo Documento',  'tipo_documento'),
            ('N° Documento',    'numero_documento'),
            ('Fecha Documento', 'fecha_documento'),
            ('Servicio',        'servicio_documento'),
            ('Fecha Desde',     'fecha_desde'),
            ('Labor',           'labor'),
            ('Modalidad',       'modalidad'),
            ('Fecha Hasta',     'fecha_hasta'),
            ('Estado Trámite',  'estado_tramite'),
        ],
        'campos': [
            {'name':'tipo_documento',   'label':'Tipo Documento',  'tipo':'text', 'base':True},
            {'name':'numero_documento', 'label':'N° Documento',    'tipo':'text', 'base':True},
            {'name':'fecha_documento',  'label':'Fecha Documento', 'tipo':'date', 'base':True},
            {'name':'servicio_documento','label':'Servicio',       'tipo':'text', 'base':True},
            {'name':'fecha_desde',      'label':'Fecha Desde',     'tipo':'date'},
            {'name':'labor',            'label':'Labor',           'tipo':'text'},
            {'name':'modalidad',        'label':'Modalidad',       'tipo':'text'},
            {'name':'fecha_hasta',      'label':'Fecha Hasta',     'tipo':'date'},
            {'name':'estado_tramite',   'label':'Estado Trámite',  'tipo':'text', 'base':True},
        ],
    },
    'licencias_medicas': {
        'nombre': 'Licencias Médicas',
        'columnas': [
            ('Tipo Documento',  'tipo_documento'),
            ('N° Documento',    'numero_documento'),
            ('Fecha Documento', 'fecha_documento'),
            ('Servicio',        'servicio_documento'),
            ('Fecha Desde',     'fecha_desde'),
            ('Fecha Hasta',     'fecha_hasta'),
            ('Días',            'dias'),
            ('Estado Trámite',  'estado_tramite'),
        ],
        'campos': [
            {'name':'tipo_documento',   'label':'Tipo Documento',  'tipo':'text',   'base':True},
            {'name':'numero_documento', 'label':'N° Documento',    'tipo':'text',   'base':True},
            {'name':'fecha_documento',  'label':'Fecha Documento', 'tipo':'date',   'base':True},
            {'name':'servicio_documento','label':'Servicio',       'tipo':'text',   'base':True},
            {'name':'fecha_desde',      'label':'Fecha Desde',     'tipo':'date'},
            {'name':'fecha_hasta',      'label':'Fecha Hasta',     'tipo':'date'},
            {'name':'dias',             'label':'Días',            'tipo':'number'},
            {'name':'estado_tramite',   'label':'Estado Trámite',  'tipo':'text',   'base':True},
        ],
    },
    'permisos_feriados': {
        'nombre': 'Permisos y Feriados',
        'columnas': [
            ('Tipo Documento',    'tipo_documento'),
            ('N° Documento',      'numero_documento'),
            ('Fecha Documento',   'fecha_documento'),
            ('Servicio',          'servicio_documento'),
            ('Fecha Desde',       'fecha_desde'),
            ('Fecha Hasta',       'fecha_hasta'),
            ('Días',              'dias'),
            ('Tipo Información',  'tipo_informacion'),
            ('Estado Trámite',    'estado_tramite'),
        ],
        'campos': [
            {'name':'tipo_documento',   'label':'Tipo Documento',   'tipo':'text', 'base':True},
            {'name':'numero_documento', 'label':'N° Documento',     'tipo':'text', 'base':True},
            {'name':'fecha_documento',  'label':'Fecha Documento',  'tipo':'date', 'base':True},
            {'name':'servicio_documento','label':'Servicio',        'tipo':'text', 'base':True},
            {'name':'fecha_desde',      'label':'Fecha Desde',      'tipo':'date'},
            {'name':'fecha_hasta',      'label':'Fecha Hasta',      'tipo':'date'},
            {'name':'dias',             'label':'Días',             'tipo':'number'},
            {'name':'tipo_informacion', 'label':'Tipo Información', 'tipo':'text'},
            {'name':'estado_tramite',   'label':'Estado Trámite',   'tipo':'text', 'base':True},
        ],
    },
    'comisiones_becas': {
        'nombre': 'Comisiones y Becas',
        'columnas': [
            ('Tipo Documento',  'tipo_documento'),
            ('N° Documento',    'numero_documento'),
            ('Fecha Documento', 'fecha_documento'),
            ('Servicio',        'servicio_documento'),
            ('Materia',         'materia'),
            ('Fecha Desde',     'fecha_desde'),
            ('Fecha Hasta',     'fecha_hasta'),
            ('Estado Trámite',  'estado_tramite'),
        ],
        'campos': [
            {'name':'tipo_documento',   'label':'Tipo Documento',  'tipo':'text', 'base':True},
            {'name':'numero_documento', 'label':'N° Documento',    'tipo':'text', 'base':True},
            {'name':'fecha_documento',  'label':'Fecha Documento', 'tipo':'date', 'base':True},
            {'name':'servicio_documento','label':'Servicio',       'tipo':'text', 'base':True},
            {'name':'materia',          'label':'Materia',         'tipo':'text'},
            {'name':'fecha_desde',      'label':'Fecha Desde',     'tipo':'date'},
            {'name':'fecha_hasta',      'label':'Fecha Hasta',     'tipo':'date'},
            {'name':'estado_tramite',   'label':'Estado Trámite',  'tipo':'text', 'base':True},
        ],
    },
    'cargas_familiares': {
        'nombre': 'Cargas Familiares',
        'columnas': [
            ('Tipo Documento',  'tipo_documento'),
            ('N° Documento',    'numero_documento'),
            ('Fecha Documento', 'fecha_documento'),
            ('Servicio',        'servicio_documento'),
            ('Tipo Carga',      'tipo_carga'),
            ('Nombre',          'nombre_carga'),
            ('Fecha Hasta',     'fecha_hasta'),
            ('Estado Trámite',  'estado_tramite'),
        ],
        'campos': [
            {'name':'tipo_documento',   'label':'Tipo Documento',  'tipo':'text', 'base':True},
            {'name':'numero_documento', 'label':'N° Documento',    'tipo':'text', 'base':True},
            {'name':'fecha_documento',  'label':'Fecha Documento', 'tipo':'date', 'base':True},
            {'name':'servicio_documento','label':'Servicio',       'tipo':'text', 'base':True},
            {'name':'tipo_carga',       'label':'Tipo Carga',      'tipo':'text'},
            {'name':'nombre_carga',     'label':'Nombre',          'tipo':'text'},
            {'name':'fecha_hasta',      'label':'Fecha Hasta',     'tipo':'date'},
            {'name':'estado_tramite',   'label':'Estado Trámite',  'tipo':'text', 'base':True},
        ],
    },
    'calificacion': {
        'nombre': 'Calificación',
        'columnas': [
            ('Tipo Documento',  'tipo_documento'),
            ('N° Documento',    'numero_documento'),
            ('Fecha Documento', 'fecha_documento'),
            ('Servicio',        'servicio_documento'),
            ('Lista',           'lista'),
            ('Puntaje',         'puntaje'),
            ('Fecha Inicio',    'fecha_inicio'),
            ('Fecha Término',   'fecha_termino'),
            ('Estado Trámite',  'estado_tramite'),
        ],
        'campos': [
            {'name':'tipo_documento',   'label':'Tipo Documento',  'tipo':'text',   'base':True},
            {'name':'numero_documento', 'label':'N° Documento',    'tipo':'text',   'base':True},
            {'name':'fecha_documento',  'label':'Fecha Documento', 'tipo':'date',   'base':True},
            {'name':'servicio_documento','label':'Servicio',       'tipo':'text',   'base':True},
            {'name':'lista',            'label':'Lista',           'tipo':'text'},
            {'name':'puntaje',          'label':'Puntaje',         'tipo':'number'},
            {'name':'fecha_inicio',     'label':'Fecha Inicio',    'tipo':'date'},
            {'name':'fecha_termino',    'label':'Fecha Término',   'tipo':'date'},
            {'name':'estado_tramite',   'label':'Estado Trámite',  'tipo':'text',   'base':True},
        ],
    },
    'destinaciones': {
        'nombre': 'Destinaciones',
        'columnas': [
            ('Tipo Documento',  'tipo_documento'),
            ('N° Documento',    'numero_documento'),
            ('Fecha Documento', 'fecha_documento'),
            ('Servicio',        'servicio_documento'),
            ('Materia Ingreso', 'materia_ingreso'),
            ('Fecha Desde',     'fecha_desde'),
            ('Estado Trámite',  'estado_tramite'),
        ],
        'campos': [
            {'name':'tipo_documento',   'label':'Tipo Documento',  'tipo':'text', 'base':True},
            {'name':'numero_documento', 'label':'N° Documento',    'tipo':'text', 'base':True},
            {'name':'fecha_documento',  'label':'Fecha Documento', 'tipo':'date', 'base':True},
            {'name':'servicio_documento','label':'Servicio',       'tipo':'text', 'base':True},
            {'name':'materia_ingreso',  'label':'Materia Ingreso', 'tipo':'text'},
            {'name':'fecha_desde',      'label':'Fecha Desde',     'tipo':'date'},
            {'name':'estado_tramite',   'label':'Estado Trámite',  'tipo':'text', 'base':True},
        ],
    },
    'ceses': {
        'nombre': 'Ceses',
        'columnas': [
            ('Tipo Documento',  'tipo_documento'),
            ('N° Documento',    'numero_documento'),
            ('Fecha Documento', 'fecha_documento'),
            ('Servicio',        'servicio_documento'),
            ('Tipo',            'tipo_cese'),
            ('Motivo',          'motivo'),
            ('Fecha Término',   'fecha_termino'),
            ('Estado Trámite',  'estado_tramite'),
        ],
        'campos': [
            {'name':'tipo_documento',   'label':'Tipo Documento',  'tipo':'text', 'base':True},
            {'name':'numero_documento', 'label':'N° Documento',    'tipo':'text', 'base':True},
            {'name':'fecha_documento',  'label':'Fecha Documento', 'tipo':'date', 'base':True},
            {'name':'servicio_documento','label':'Servicio',       'tipo':'text', 'base':True},
            {'name':'tipo_cese',        'label':'Tipo',            'tipo':'text'},
            {'name':'motivo',           'label':'Motivo',          'tipo':'text'},
            {'name':'fecha_termino',    'label':'Fecha Término',   'tipo':'date'},
            {'name':'estado_tramite',   'label':'Estado Trámite',  'tipo':'text', 'base':True},
        ],
    },
    'anotaciones': {
        'nombre': 'Anotaciones',
        'columnas': [
            ('Tipo Documento',    'tipo_documento'),
            ('N° Documento',      'numero_documento'),
            ('Fecha Documento',   'fecha_documento'),
            ('Servicio',          'servicio_documento'),
            ('Tipo',              'tipo_anotacion'),
            ('Fecha',             'fecha_anotacion'),
            ('Run Responsable',   'run_responsable'),
            ('Nombre Responsable','nombre_responsable'),
            ('Estado Trámite',    'estado_tramite'),
        ],
        'campos': [
            {'name':'tipo_documento',    'label':'Tipo Documento',     'tipo':'text', 'base':True},
            {'name':'numero_documento',  'label':'N° Documento',       'tipo':'text', 'base':True},
            {'name':'fecha_documento',   'label':'Fecha Documento',    'tipo':'date', 'base':True},
            {'name':'servicio_documento','label':'Servicio',           'tipo':'text', 'base':True},
            {'name':'tipo_anotacion',    'label':'Tipo',               'tipo':'text'},
            {'name':'fecha_anotacion',   'label':'Fecha',              'tipo':'date'},
            {'name':'run_responsable',   'label':'Run Responsable',    'tipo':'text'},
            {'name':'nombre_responsable','label':'Nombre Responsable', 'tipo':'text'},
            {'name':'estado_tramite',    'label':'Estado Trámite',     'tipo':'text', 'base':True},
        ],
    },
    'declaraciones_patrimonio': {
        'nombre': 'Declaraciones de Patrimonio y/o Intereses',
        'columnas': [
            ('Tipo Documento',     'tipo_documento'),
            ('N° Documento',       'numero_documento'),
            ('Fecha Documento',    'fecha_documento'),
            ('Servicio',           'servicio_documento'),
            ('Materia Declaración','materia_declaracion'),
            ('Fecha Declaración',  'fecha_declaracion'),
            ('Fecha Próxima',      'fecha_proxima'),
            ('Estado Trámite',     'estado_tramite'),
        ],
        'campos': [
            {'name':'tipo_documento',    'label':'Tipo Documento',      'tipo':'text', 'base':True},
            {'name':'numero_documento',  'label':'N° Documento',        'tipo':'text', 'base':True},
            {'name':'fecha_documento',   'label':'Fecha Documento',     'tipo':'date', 'base':True},
            {'name':'servicio_documento','label':'Servicio',            'tipo':'text', 'base':True},
            {'name':'materia_declaracion','label':'Materia Declaración','tipo':'text'},
            {'name':'fecha_declaracion', 'label':'Fecha Declaración',   'tipo':'date'},
            {'name':'fecha_proxima',     'label':'Fecha Próxima',       'tipo':'date'},
            {'name':'estado_tramite',    'label':'Estado Trámite',      'tipo':'text', 'base':True},
        ],
    },
    'modifica_rectifica': {
        'nombre': 'Modifica / Rectifica',
        'columnas': [
            ('Tipo Documento',  'tipo_documento'),
            ('N° Documento',    'numero_documento'),
            ('Fecha Documento', 'fecha_documento'),
            ('Servicio',        'servicio_documento'),
            ('Materia',         'materia_documento'),
            ('Modifica a',      'modifica_a'),
            ('Documento N',     'documento_n'),
            ('Materia Relac.',  'materia_relacionada'),
            ('Estado Trámite',  'estado_tramite'),
        ],
        'campos': [
            {'name':'tipo_documento',    'label':'Tipo Documento',    'tipo':'text', 'base':True},
            {'name':'numero_documento',  'label':'N° Documento',      'tipo':'text', 'base':True},
            {'name':'fecha_documento',   'label':'Fecha Documento',   'tipo':'date', 'base':True},
            {'name':'servicio_documento','label':'Servicio',          'tipo':'text', 'base':True},
            {'name':'materia_documento', 'label':'Materia Documento', 'tipo':'text'},
            {'name':'modifica_a',        'label':'Modifica Rectifica a','tipo':'text'},
            {'name':'documento_n',       'label':'Documento N',       'tipo':'text'},
            {'name':'materia_relacionada','label':'Materia Relacionada','tipo':'text'},
            {'name':'estado_tramite',    'label':'Estado Trámite',    'tipo':'text', 'base':True},
        ],
    },
}

CAMPOS_BASE = {'tipo_documento', 'numero_documento', 'fecha_documento', 'servicio_documento', 'estado_tramite'}

# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if not session.get('usuario_id'):
            return redirect(url_for('login', next=request.path))
        return f(*a, **kw)
    return dec

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute('PRAGMA foreign_keys = ON')
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db: db.close()

def ext_ok(filename, exts):
    return '.' in filename and filename.rsplit('.',1)[1].lower() in exts

def guardar_archivos(campo, exts):
    guardados = []
    for f in request.files.getlist(campo):
        if f and f.filename and ext_ok(f.filename, exts):
            ext = f.filename.rsplit('.',1)[1].lower()
            nombre = f'{uuid.uuid4().hex}.{ext}'
            f.save(os.path.join(UPLOADS, nombre))
            guardados.append((nombre, f.filename))
    return guardados

def borrar_archivo(nombre):
    if not nombre: return
    ruta = os.path.join(UPLOADS, nombre)
    if os.path.exists(ruta): os.remove(ruta)

def guardar_foto(campo):
    f = request.files.get(campo)
    if f and f.filename and ext_ok(f.filename, EXT_IMG):
        ext = f.filename.rsplit('.',1)[1].lower()
        nombre = f'{uuid.uuid4().hex}.{ext}'
        f.save(os.path.join(UPLOADS, nombre))
        return nombre
    return None

def rut_valido(rut):
    limpio = rut.upper().replace('.','').replace('-','').strip()
    if len(limpio) < 2: return False
    cuerpo, dv = limpio[:-1], limpio[-1]
    if not cuerpo.isdigit(): return False
    s, m = 0, 2
    for d in reversed(cuerpo):
        s += int(d)*m; m = 2 if m==7 else m+1
    esperado = {11:'0', 10:'K'}.get(11-(s%11), str(11-(s%11)))
    return dv == esperado

def fechas_ok(desde, hasta):
    if desde and hasta:
        try: return date.fromisoformat(desde) <= date.fromisoformat(hasta)
        except ValueError: return True
    return True

def auditoria(accion, entidad, eid, desc=''):
    db = get_db()
    db.execute('INSERT INTO auditoria (usuario_id,usuario_nombre,accion,entidad,entidad_id,descripcion) VALUES(?,?,?,?,?,?)',
               (session.get('usuario_id'), session.get('usuario_nombre',''), accion, entidad, eid, desc))
    db.commit()

# ─────────────────────────────────────────────
#  INIT DB
# ─────────────────────────────────────────────
def init_db():
    db = sqlite3.connect(DB_PATH)
    db.execute('PRAGMA foreign_keys = ON')

    db.execute('''CREATE TABLE IF NOT EXISTS fichas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        numero_ficha INTEGER,
        nombre TEXT NOT NULL,
        rut TEXT NOT NULL UNIQUE,
        rol TEXT, pasaporte TEXT, pais_nacionalidad TEXT,
        region TEXT, comuna TEXT, ciudad TEXT, correo TEXT,
        titulo TEXT, grado_academico TEXT, jerarquia TEXT,
        direccion TEXT, observaciones TEXT,
        fecha_nacimiento TEXT, estado_civil TEXT, telefono TEXT, sexo TEXT, link TEXT,
        foto TEXT,
        activo INTEGER NOT NULL DEFAULT 1,
        fecha_creacion TEXT DEFAULT CURRENT_TIMESTAMP
    )''')

    db.execute('''CREATE TABLE IF NOT EXISTS documentos_hv (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ficha_id INTEGER NOT NULL,
        seccion TEXT NOT NULL,
        tipo_documento TEXT,
        numero_documento TEXT,
        fecha_documento TEXT,
        servicio_documento TEXT,
        estado_tramite TEXT,
        datos_extra TEXT DEFAULT '{}',
        fecha_creacion TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (ficha_id) REFERENCES fichas(id) ON DELETE CASCADE
    )''')

    db.execute('''CREATE TABLE IF NOT EXISTS archivos_documentos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        documento_id INTEGER NOT NULL,
        nombre_archivo TEXT NOT NULL,
        nombre_original TEXT,
        fecha_creacion TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (documento_id) REFERENCES documentos_hv(id) ON DELETE CASCADE
    )''')

    db.execute('''CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL UNIQUE, password_hash TEXT NOT NULL,
        nombre TEXT, activo INTEGER NOT NULL DEFAULT 1,
        fecha_creacion TEXT DEFAULT CURRENT_TIMESTAMP
    )''')

    db.execute('''CREATE TABLE IF NOT EXISTS auditoria (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario_id INTEGER, usuario_nombre TEXT,
        accion TEXT, entidad TEXT, entidad_id INTEGER, descripcion TEXT,
        fecha_creacion TEXT DEFAULT CURRENT_TIMESTAMP
    )''')

    db.execute('''CREATE TABLE IF NOT EXISTS catalogos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        categoria TEXT NOT NULL, nombre TEXT NOT NULL,
        UNIQUE(categoria, nombre)
    )''')

    if db.execute('SELECT COUNT(*) FROM usuarios').fetchone()[0] == 0:
        db.execute('INSERT INTO usuarios (username,password_hash,nombre,activo) VALUES(?,?,?,1)',
                   (ADMIN_USER, generate_password_hash(ADMIN_PASSWORD), 'Administrador'))

    semillas = {
        'contrato':  ['Planta','Contrata','Honorarios','Suplencia','Interino'],
        'estamento': ['Directivo','Profesional','Técnico','Administrativo','Auxiliar','Docente'],
        'jornada':   ['Completa (44h)','Media jornada (22h)','Parcial','Otra'],
        'estado_tramite': ['Registrado - Válido','Registrado - Modificado','Pendiente','Anulado'],
        'tipo_documento': ['Resolución Exenta','Decreto','Oficio','Resolución'],
    }
    for cat, vals in semillas.items():
        if db.execute('SELECT COUNT(*) FROM catalogos WHERE categoria=?',(cat,)).fetchone()[0] == 0:
            for v in vals:
                db.execute('INSERT INTO catalogos (categoria,nombre) VALUES(?,?)',(cat,v))

    db.commit(); db.close()

# ─────────────────────────────────────────────
#  LOGIN / LOGOUT
# ─────────────────────────────────────────────
@app.route('/')
def home(): return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET','POST'])
def login():
    if session.get('usuario_id'): return redirect(url_for('dashboard'))
    if request.method == 'POST':
        u = request.form.get('usuario','').strip()
        c = request.form.get('clave','').strip()
        db = get_db()
        row = db.execute('SELECT * FROM usuarios WHERE username=? AND activo=1',(u,)).fetchone()
        if row and check_password_hash(row['password_hash'], c):
            session['usuario_id']     = row['id']
            session['usuario_nombre'] = row['nombre'] or row['username']
            return redirect(request.args.get('next') or url_for('dashboard'))
        flash('Usuario o contraseña incorrectos.','error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ─────────────────────────────────────────────
#  DASHBOARD
# ─────────────────────────────────────────────
ESTADOS_VALIDOS = ('registrado - válido', 'registrado - valido', 'válido', 'valido', 'registrado')

@app.route('/dashboard')
@login_required
def dashboard():
    db = get_db()

    total_fichas = db.execute('SELECT COUNT(*) FROM fichas WHERE activo=1').fetchone()[0]
    en_papelera  = db.execute('SELECT COUNT(*) FROM fichas WHERE activo=0').fetchone()[0]

    # Solo documentos con estado válido/registrado
    docs_validos = db.execute("""
        SELECT COUNT(*) FROM documentos_hv
        WHERE LOWER(TRIM(estado_tramite)) IN ({})
    """.format(','.join('?'*len(ESTADOS_VALIDOS))), ESTADOS_VALIDOS).fetchone()[0]

    # Documentos válidos por sección (top 8)
    por_seccion = db.execute("""
        SELECT seccion, COUNT(*) as total FROM documentos_hv
        WHERE LOWER(TRIM(estado_tramite)) IN ({})
        GROUP BY seccion ORDER BY total DESC LIMIT 8
    """.format(','.join('?'*len(ESTADOS_VALIDOS))), ESTADOS_VALIDOS).fetchall()

    # Designaciones vigentes (fecha_hasta vacía o futura) y válidas
    from datetime import date
    hoy = date.today().isoformat()
    vigentes = db.execute("""
        SELECT COUNT(*) FROM documentos_hv
        WHERE seccion='designaciones'
        AND LOWER(TRIM(estado_tramite)) IN ({})
        AND (
            json_extract(datos_extra,'$.fecha_hasta') = ''
            OR json_extract(datos_extra,'$.fecha_hasta') IS NULL
            OR json_extract(datos_extra,'$.fecha_hasta') >= ?
        )
    """.format(','.join('?'*len(ESTADOS_VALIDOS))), ESTADOS_VALIDOS + (hoy,)).fetchone()[0]

    # Contratos que vencen en los próximos 60 días
    from datetime import timedelta
    en_60 = (date.today() + timedelta(days=60)).isoformat()
    proximos_vencer = db.execute("""
        SELECT f.nombre, f.rut, json_extract(d.datos_extra,'$.fecha_hasta') as vence,
               d.servicio_documento, json_extract(d.datos_extra,'$.calidad') as calidad
        FROM documentos_hv d
        JOIN fichas f ON f.id = d.ficha_id
        WHERE d.seccion = 'designaciones'
        AND LOWER(TRIM(d.estado_tramite)) IN ({})
        AND json_extract(d.datos_extra,'$.fecha_hasta') BETWEEN ? AND ?
        ORDER BY vence ASC LIMIT 10
    """.format(','.join('?'*len(ESTADOS_VALIDOS))), ESTADOS_VALIDOS + (hoy, en_60)).fetchall()

    # Últimas 5 fichas creadas
    ultimas_fichas = db.execute("""
        SELECT nombre, rut, fecha_creacion FROM fichas
        WHERE activo=1 ORDER BY fecha_creacion DESC LIMIT 5
    """).fetchall()

    nombres_sec = {k: v['nombre'] for k, v in SECCIONES_CONFIG.items()}

    return render_template('dashboard.html',
        total_fichas=total_fichas, en_papelera=en_papelera,
        docs_validos=docs_validos, vigentes=vigentes,
        por_seccion=por_seccion, proximos_vencer=proximos_vencer,
        ultimas_fichas=ultimas_fichas, nombres_sec=nombres_sec)

# ─────────────────────────────────────────────
#  LISTADO
# ─────────────────────────────────────────────
@app.route('/fichas')
@login_required
def fichas_listado():
    q         = request.args.get('q','').strip()
    unidad    = request.args.get('unidad','').strip()
    contrato  = request.args.get('contrato','').strip()
    db        = get_db()

    sql = '''SELECT f.*, COUNT(DISTINCT d.id) AS total_docs,
                    MAX(CASE WHEN d.seccion='designaciones' THEN json_extract(d.datos_extra,'$.calidad') END) AS calidad_actual,
                    MAX(CASE WHEN d.seccion='designaciones' THEN d.servicio_documento END) AS unidad_actual
             FROM fichas f
             LEFT JOIN documentos_hv d ON d.ficha_id=f.id
             WHERE f.activo=1'''
    params = []
    if q:
        sql += ' AND (f.nombre LIKE ? OR f.rut LIKE ?)'; params += [f'%{q}%', f'%{q}%']

    sql += ' GROUP BY f.id ORDER BY f.nombre COLLATE NOCASE'
    fichas = db.execute(sql, params).fetchall()
    en_papelera = db.execute('SELECT COUNT(*) FROM fichas WHERE activo=0').fetchone()[0]
    return render_template('listado.html', fichas=fichas, q=q, en_papelera=en_papelera)

@app.route('/fichas/<int:fid>/pdf')
@login_required
def ficha_pdf(fid):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

    db = get_db()
    ficha = db.execute('SELECT * FROM fichas WHERE id=?', (fid,)).fetchone()
    if not ficha:
        flash('Ficha no encontrada.', 'error')
        return redirect(url_for('fichas_listado'))

    # Solo documentos con estado válido
    documentos = {}
    for sec_id in SECCIONES_CONFIG:
        docs = db.execute("""
            SELECT * FROM documentos_hv
            WHERE ficha_id=? AND seccion=?
            AND LOWER(TRIM(estado_tramite)) IN ({})
            ORDER BY fecha_documento DESC
        """.format(','.join('?'*len(ESTADOS_VALIDOS))),
        (fid, sec_id) + ESTADOS_VALIDOS).fetchall()
        if docs:
            documentos[sec_id] = [dict(d) for d in docs]
            for d in documentos[sec_id]:
                try:
                    d['extra'] = json.loads(d['datos_extra'] or '{}')
                except Exception:
                    d['extra'] = {}

    buf = BytesIO()
    pdf_doc = SimpleDocTemplate(buf, pagesize=letter,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    COLOR_PRIMARIO = colors.HexColor('#1C3D5A')
    COLOR_GRIS     = colors.HexColor('#F5F7FA')
    COLOR_BORDE    = colors.HexColor('#DDE1E7')
    COLOR_ACENTO   = colors.HexColor('#9C2B2B')

    estilos = getSampleStyleSheet()
    est_titulo = ParagraphStyle('titulo', parent=estilos['Normal'],
                                fontSize=15, fontName='Helvetica-Bold',
                                textColor=colors.white, alignment=TA_CENTER)
    est_sub    = ParagraphStyle('sub', parent=estilos['Normal'],
                                fontSize=8, fontName='Helvetica',
                                textColor=colors.HexColor('#BFD4E8'), alignment=TA_CENTER)
    est_etq    = ParagraphStyle('etq', parent=estilos['Normal'],
                                fontSize=8, fontName='Helvetica-Bold',
                                textColor=colors.HexColor('#5B6470'), alignment=TA_RIGHT)
    est_val    = ParagraphStyle('val', parent=estilos['Normal'],
                                fontSize=9, fontName='Helvetica',
                                textColor=colors.HexColor('#161A1F'))
    est_seccion= ParagraphStyle('sec', parent=estilos['Normal'],
                                fontSize=8, fontName='Helvetica-Bold',
                                textColor=COLOR_PRIMARIO, spaceAfter=4)
    est_tabla_enc = ParagraphStyle('tenc', parent=estilos['Normal'],
                                   fontSize=7, fontName='Helvetica-Bold',
                                   textColor=colors.HexColor('#5B6470'))
    est_tabla_cel = ParagraphStyle('tcel', parent=estilos['Normal'],
                                   fontSize=7.5, fontName='Helvetica',
                                   textColor=colors.HexColor('#161A1F'))

    story = []

    # ── Encabezado ──
    enc_data = [[Paragraph(ficha['nombre'] or '', est_titulo)],
                [Paragraph('Hoja de Vida de Personal', est_sub)]]
    enc_table = Table(enc_data, colWidths=[17*cm])
    enc_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), COLOR_PRIMARIO),
        ('ROWPADDING', (0,0), (-1,-1), 8),
        ('TOPPADDING', (0,0), (-1,0), 14),
        ('BOTTOMPADDING', (0,-1), (-1,-1), 14),
        ('ROUNDEDCORNERS', [6,6,0,0]),
    ]))
    story.append(enc_table)
    story.append(Spacer(1, 0.3*cm))

    # ── Datos personales ──
    def campo(etq, val):
        return [Paragraph(etq, est_etq), Paragraph(str(val or '—'), est_val)]

    dp_data = [
        campo('Run (RUT)',           ficha['rut']),
        campo('ROL',                 ficha['rol']),
        campo('Pasaporte',           ficha['pasaporte']),
        campo('País Nacionalidad',   ficha['pais_nacionalidad']),
        campo('Nombre Completo',     ficha['nombre']),
        campo('Correo',              ficha['correo']),
        campo('Región',              ficha['region']),
        campo('Comuna',              ficha['comuna']),
        campo('Ciudad',              ficha['ciudad']),
        campo('Fecha Nacimiento',    ficha['fecha_nacimiento']),
        campo('Estado Civil',        ficha['estado_civil']),
        campo('Jerarquía',           ficha['jerarquia']),
        campo('Grado Académico',     ficha['grado_academico']),
        campo('Título',              ficha['titulo']),
        campo('Teléfono',            ficha['telefono']),
        campo('Dirección',           ficha['direccion']),
    ]
    if ficha['observaciones']:
        dp_data.append(campo('Observaciones', ficha['observaciones']))

    dp_table = Table(dp_data, colWidths=[4.5*cm, 12.5*cm])
    dp_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), COLOR_GRIS),
        ('BOX', (0,0), (-1,-1), 0.5, COLOR_BORDE),
        ('INNERGRID', (0,0), (-1,-1), 0.3, COLOR_BORDE),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (0,-1), 6),
        ('RIGHTPADDING', (0,0), (0,-1), 6),
    ]))
    story.append(dp_table)

    # ── Secciones con documentos válidos ──
    for sec_id, sec_conf in SECCIONES_CONFIG.items():
        docs = documentos.get(sec_id)
        if not docs:
            continue

        story.append(Spacer(1, 0.5*cm))
        story.append(HRFlowable(width='100%', thickness=0.5, color=COLOR_BORDE))
        story.append(Spacer(1, 0.2*cm))
        story.append(Paragraph(sec_conf['nombre'].upper(), est_seccion))

        # Construir columnas de la tabla
        cols_sec = sec_conf['columnas']
        enc_fila = [Paragraph(h, est_tabla_enc) for h, _ in cols_sec]

        filas_tabla = [enc_fila]
        for doc in docs:
            fila = []
            for _, clave in cols_sec:
                if clave in ('tipo_documento','numero_documento','fecha_documento',
                             'servicio_documento','estado_tramite'):
                    fila.append(Paragraph(str(doc.get(clave,'') or ''), est_tabla_cel))
                else:
                    fila.append(Paragraph(str(doc['extra'].get(clave,'') or ''), est_tabla_cel))
            filas_tabla.append(fila)

        n_cols = len(cols_sec)
        ancho_col = 17*cm / n_cols
        t = Table(filas_tabla, colWidths=[ancho_col]*n_cols, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), COLOR_PRIMARIO),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('BACKGROUND', (0,1), (-1,-1), colors.white),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, COLOR_GRIS]),
            ('BOX', (0,0), (-1,-1), 0.5, COLOR_BORDE),
            ('INNERGRID', (0,0), (-1,-1), 0.3, COLOR_BORDE),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('FONTSIZE', (0,0), (-1,-1), 7),
            ('WORDWRAP', (0,0), (-1,-1), True),
        ]))
        story.append(t)

    # ── Pie de página ──
    story.append(Spacer(1, 0.8*cm))
    from datetime import date as dt
    pie = Paragraph(
        f'Documento generado el {dt.today().strftime("%d/%m/%Y")} — Solo incluye documentos con estado Registrado - Válido',
        ParagraphStyle('pie', parent=estilos['Normal'], fontSize=7,
                       textColor=colors.HexColor('#9B9B9B'), alignment=TA_CENTER))
    story.append(pie)

    pdf_doc.build(story)
    buf.seek(0)
    nombre_archivo = f"hoja_vida_{(ficha['nombre'] or 'persona').replace(' ','_')}.pdf"
    return send_file(buf, as_attachment=True, download_name=nombre_archivo,
                     mimetype='application/pdf')


@app.route('/fichas/exportar-fichas')
@login_required
def fichas_exportar_fichas():
    db = get_db()
    fichas = db.execute('SELECT * FROM fichas WHERE activo=1 ORDER BY nombre COLLATE NOCASE').fetchall()
    wb = Workbook(); ws = wb.active; ws.title = 'Hojas de Vida'

    enc = ['Run (RUT)','ROL','Pasaporte','País Nacionalidad','Nombre Completo','Correo',
           'Región','Comuna','Ciudad','Fecha Nacimiento','Estado Civil',
           'Jerarquía','Grado Académico','Título','Teléfono','Dirección','Observaciones']
    ws.append(enc)
    for c in ws[1]:
        c.font = Font(name='Arial', bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', start_color='1C3D5A')
        c.alignment = Alignment(horizontal='left')

    for f in fichas:
        ws.append([f['rut'], f['rol'], f['pasaporte'], f['pais_nacionalidad'],
                   f['nombre'], f['correo'], f['region'], f['comuna'], f['ciudad'],
                   f['fecha_nacimiento'], f['estado_civil'], f['jerarquia'],
                   f['grado_academico'], f['titulo'], f['telefono'],
                   f['direccion'], f['observaciones']])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            max(len(str(c.value or '')) for c in col) + 3, 12)

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='hojas_de_vida.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/fichas/exportar-actos')
@login_required
def fichas_exportar_actos():
    db  = get_db()
    wb  = Workbook()
    primera = True

    # nombres de display para cada sección
    nombres_display = {
        'designaciones':                  'Designaciones',
        'cese_funciones':                 'Cese de Funciones',
        'responsabilidad_administrativa': 'Responsabilidad Administrativa',
        'inhabilidades':                  'Inhabilidades',
        'estudios':                       'Estudios',
        'contrato_honorarios':            'Contrato a Honorarios',
        'licencias_medicas':              'Licencias Médicas',
        'permisos_feriados':              'Permisos y Feriados',
        'comisiones_becas':               'Comisiones y Becas',
        'cargas_familiares':              'Cargas Familiares',
        'calificacion':                   'Calificación',
        'destinaciones':                  'Destinaciones',
        'ceses':                          'Ceses',
        'anotaciones':                    'Anotaciones',
        'declaraciones_patrimonio':       'Declaraciones de Patrimonio',
        'modifica_rectifica':             'Modifica Rectifica',
    }

    # obtener secciones que tienen datos
    secciones_con_datos = [r[0] for r in db.execute(
        'SELECT DISTINCT seccion FROM documentos_hv ORDER BY seccion').fetchall()]

    if not secciones_con_datos:
        # crear hoja vacía si no hay datos
        ws = wb.active; ws.title = 'Sin datos'
        ws.append(['No hay actos administrativos registrados.'])
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, as_attachment=True, download_name='actos_administrativos.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    for seccion in secciones_con_datos:
        nombre_hoja = nombres_display.get(seccion, seccion)[:31]  # Excel limita a 31 chars

        docs = db.execute('''
            SELECT f.rut, f.nombre, d.*
            FROM documentos_hv d
            JOIN fichas f ON f.id = d.ficha_id
            WHERE d.seccion = ?
            ORDER BY f.nombre COLLATE NOCASE, d.fecha_documento
        ''', (seccion,)).fetchall()

        if not docs: continue

        # construir encabezados dinámicos: RUT + Nombre + campos base + claves de datos_extra
        claves_extra = set()
        for doc in docs:
            try:
                extra = json.loads(doc['datos_extra'] or '{}')
                claves_extra.update(extra.keys())
            except Exception:
                pass
        claves_extra = sorted(claves_extra)

        enc = ['RUT', 'Nombre', 'Tipo Documento', 'N° Documento', 'Fecha Documento',
               'Servicio Documento', 'Estado Trámite'] + [k.replace('_',' ').title() for k in claves_extra]

        if primera:
            ws = wb.active; ws.title = nombre_hoja; primera = False
        else:
            ws = wb.create_sheet(nombre_hoja)

        ws.append(enc)
        for c in ws[1]:
            c.font = Font(name='Arial', bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', start_color='1C3D5A')
            c.alignment = Alignment(horizontal='left')

        for doc in docs:
            try:
                extra = json.loads(doc['datos_extra'] or '{}')
            except Exception:
                extra = {}
            fila = [doc['rut'], doc['nombre'],
                    doc['tipo_documento'], doc['numero_documento'],
                    doc['fecha_documento'], doc['servicio_documento'], doc['estado_tramite']]
            fila += [extra.get(k, '') for k in claves_extra]
            ws.append(fila)

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                max(len(str(c.value or '')) for c in col) + 3, 12)

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='actos_administrativos.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/fichas/papelera')
@login_required
def fichas_papelera():
    db = get_db()
    fichas = db.execute('SELECT * FROM fichas WHERE activo=0 ORDER BY nombre COLLATE NOCASE').fetchall()
    return render_template('papelera.html', fichas=fichas)

@app.route('/fichas/<int:fid>/restaurar', methods=['POST'])
@login_required
def ficha_restaurar(fid):
    db = get_db(); db.execute('UPDATE fichas SET activo=1 WHERE id=?',(fid,)); db.commit()
    flash('Ficha restaurada.','success')
    return redirect(url_for('fichas_papelera'))

@app.route('/fichas/<int:fid>/eliminar_definitivo', methods=['POST'])
@login_required
def ficha_eliminar_definitivo(fid):
    db = get_db(); db.execute('DELETE FROM fichas WHERE id=?',(fid,)); db.commit()
    flash('Ficha eliminada definitivamente.','success')
    return redirect(url_for('fichas_papelera'))

# ─────────────────────────────────────────────
#  FICHA — CREAR / VER / EDITAR
# ─────────────────────────────────────────────
def _leer_ficha():
    g = lambda k: request.form.get(k,'').strip()
    return {k: g(k) for k in ['nombre','rut','rol','pasaporte','pais_nacionalidad',
            'region','comuna','ciudad','correo','titulo','grado_academico','jerarquia',
            'direccion','observaciones','fecha_nacimiento','estado_civil','telefono','sexo','link']}

@app.route('/fichas/nueva', methods=['GET','POST'])
@login_required
def ficha_nueva():
    db = get_db()
    if request.method == 'POST':
        d = _leer_ficha()
        if not d['nombre'] or not d['rut']:
            flash('Nombre y RUT son obligatorios.','error'); return redirect(url_for('ficha_nueva'))
        if not rut_valido(d['rut']):
            flash('RUT inválido (revisa el dígito verificador).','error'); return redirect(url_for('ficha_nueva'))
        if db.execute('SELECT id FROM fichas WHERE rut=?',(d['rut'],)).fetchone():
            flash('Ya existe una ficha con ese RUT.','error'); return redirect(url_for('ficha_nueva'))
        foto = guardar_foto('foto')

        # Calcular el menor número correlativo disponible (rellena huecos)
        numeros_usados = {r[0] for r in db.execute('SELECT numero_ficha FROM fichas WHERE numero_ficha IS NOT NULL')}
        numero_ficha = 1
        while numero_ficha in numeros_usados:
            numero_ficha += 1

        cur  = db.execute('''INSERT INTO fichas (numero_ficha,nombre,rut,rol,pasaporte,pais_nacionalidad,
                             region,comuna,ciudad,correo,titulo,grado_academico,jerarquia,
                             direccion,observaciones,fecha_nacimiento,estado_civil,telefono,sexo,link,foto)
                             VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                          (numero_ficha,) + tuple(d[k] for k in ['nombre','rut','rol','pasaporte','pais_nacionalidad',
                                'region','comuna','ciudad','correo','titulo','grado_academico','jerarquia',
                                'direccion','observaciones','fecha_nacimiento','estado_civil','telefono','sexo','link']) + (foto,))
        db.commit()
        auditoria('crear','ficha',cur.lastrowid,f"Ficha creada: {d['nombre']} ({d['rut']})")
        flash('Ficha creada.','success')
        return redirect(url_for('ficha_detalle', fid=cur.lastrowid))

    cats = _cargar_catalogos(db)
    return render_template('ficha.html', ficha=None, documentos={},
                           secciones_config=SECCIONES_CONFIG, cats=cats,
                           secciones_json=json.dumps({k:v for k,v in SECCIONES_CONFIG.items()}))

def _cargar_catalogos(db):
    cats = {}
    for cat in ['contrato','estamento','jornada','estado_tramite','tipo_documento']:
        cats[cat] = [r['nombre'] for r in db.execute('SELECT nombre FROM catalogos WHERE categoria=? ORDER BY nombre',(cat,))]
    return cats

@app.route('/fichas/<int:fid>')
@login_required
def ficha_detalle(fid):
    db = get_db()
    ficha = db.execute('SELECT * FROM fichas WHERE id=?',(fid,)).fetchone()
    if not ficha:
        flash('Ficha no encontrada.','error'); return redirect(url_for('fichas_listado'))

    # cargar todos los documentos agrupados por sección
    documentos = {s:[] for s in SECCIONES_CONFIG}
    for row in db.execute('SELECT * FROM documentos_hv WHERE ficha_id=? ORDER BY id DESC',(fid,)):
        sec = row['seccion']
        if sec not in documentos: continue
        d = dict(row)
        d['datos_extra'] = json.loads(d['datos_extra'] or '{}')
        archs = db.execute('SELECT * FROM archivos_documentos WHERE documento_id=? ORDER BY id',(d['id'],)).fetchall()
        d['archivos'] = archs
        d['archivos_json'] = json.dumps([{
            'id':a['id'],'nombre':a['nombre_original'] or a['nombre_archivo'],
            'url':url_for('static', filename='uploads/'+a['nombre_archivo'])
        } for a in archs])
        documentos[sec].append(d)

    cats = _cargar_catalogos(db)
    return render_template('ficha.html', ficha=ficha, documentos=documentos,
                           secciones_config=SECCIONES_CONFIG, cats=cats,
                           secciones_json=json.dumps({k:{'nombre':v['nombre'],'campos':v['campos']} for k,v in SECCIONES_CONFIG.items()}))

@app.route('/fichas/<int:fid>/actualizar', methods=['POST'])
@login_required
def ficha_actualizar(fid):
    db = get_db()
    actual = db.execute('SELECT * FROM fichas WHERE id=?',(fid,)).fetchone()
    if not actual: flash('Ficha no encontrada.','error'); return redirect(url_for('fichas_listado'))
    d = _leer_ficha()
    if not d['nombre'] or not d['rut']:
        flash('Nombre y RUT son obligatorios.','error'); return redirect(url_for('ficha_detalle', fid=fid))
    if not rut_valido(d['rut']):
        flash('RUT inválido.','error'); return redirect(url_for('ficha_detalle', fid=fid))
    if db.execute('SELECT id FROM fichas WHERE rut=? AND id!=?',(d['rut'],fid)).fetchone():
        flash('RUT ya existe en otra ficha.','error'); return redirect(url_for('ficha_detalle', fid=fid))
    foto = guardar_foto('foto') or actual['foto']
    db.execute('''UPDATE fichas SET nombre=?,rut=?,rol=?,pasaporte=?,pais_nacionalidad=?,
                  region=?,comuna=?,ciudad=?,correo=?,titulo=?,grado_academico=?,jerarquia=?,
                  direccion=?,observaciones=?,fecha_nacimiento=?,estado_civil=?,telefono=?,sexo=?,link=?,foto=?
                  WHERE id=?''',
               tuple(d[k] for k in ['nombre','rut','rol','pasaporte','pais_nacionalidad',
                     'region','comuna','ciudad','correo','titulo','grado_academico','jerarquia',
                     'direccion','observaciones','fecha_nacimiento','estado_civil','telefono','sexo','link']) + (foto,fid))
    db.commit()
    auditoria('actualizar','ficha',fid,f"Ficha actualizada: {d['nombre']}")
    flash('Ficha actualizada.','success')
    return redirect(url_for('ficha_detalle', fid=fid))

@app.route('/fichas/<int:fid>/eliminar', methods=['POST'])
@login_required
def ficha_eliminar(fid):
    db = get_db()
    db.execute('UPDATE fichas SET activo=0 WHERE id=?',(fid,)); db.commit()
    flash('Ficha movida a la papelera.','success')
    return redirect(url_for('fichas_listado'))

# ─────────────────────────────────────────────
#  DOCUMENTOS HV (genérico para todas las secciones)
# ─────────────────────────────────────────────
def _leer_doc(seccion):
    g = lambda k: request.form.get(k,'').strip()
    base = {k: g(k) for k in CAMPOS_BASE}
    config = SECCIONES_CONFIG[seccion]
    extra = {}
    for campo in config['campos']:
        if campo.get('base'): continue
        extra[campo['name']] = g(campo['name'])
    return base, extra

@app.route('/fichas/<int:fid>/doc/<seccion>/guardar', methods=['POST'])
@login_required
def doc_guardar(fid, seccion):
    if seccion not in SECCIONES_CONFIG:
        flash('Sección inválida.','error'); return redirect(url_for('ficha_detalle', fid=fid))
    base, extra = _leer_doc(seccion)
    db = get_db()
    cur = db.execute('''INSERT INTO documentos_hv (ficha_id,seccion,tipo_documento,numero_documento,
                        fecha_documento,servicio_documento,estado_tramite,datos_extra)
                        VALUES(?,?,?,?,?,?,?,?)''',
                     (fid, seccion, base['tipo_documento'], base['numero_documento'],
                      base['fecha_documento'], base['servicio_documento'],
                      base['estado_tramite'], json.dumps(extra)))
    doc_id = cur.lastrowid
    for nombre, original in guardar_archivos('archivos', EXT_DOC):
        db.execute('INSERT INTO archivos_documentos (documento_id,nombre_archivo,nombre_original) VALUES(?,?,?)',
                   (doc_id, nombre, original))
    db.commit()
    auditoria('crear', seccion, doc_id, f"{base['tipo_documento']} {base['numero_documento']}")
    flash('Registro guardado.','success')
    return redirect(url_for('ficha_detalle', fid=fid))

@app.route('/fichas/<int:fid>/doc/<seccion>/modificar', methods=['POST'])
@login_required
def doc_modificar(fid, seccion):
    if seccion not in SECCIONES_CONFIG:
        flash('Sección inválida.','error'); return redirect(url_for('ficha_detalle', fid=fid))
    doc_id = request.form.get('doc_id','').strip()
    if not doc_id:
        flash('Selecciona un registro.','error'); return redirect(url_for('ficha_detalle', fid=fid))
    base, extra = _leer_doc(seccion)
    db = get_db()
    db.execute('''UPDATE documentos_hv SET tipo_documento=?,numero_documento=?,fecha_documento=?,
                  servicio_documento=?,estado_tramite=?,datos_extra=?
                  WHERE id=? AND ficha_id=?''',
               (base['tipo_documento'], base['numero_documento'], base['fecha_documento'],
                base['servicio_documento'], base['estado_tramite'], json.dumps(extra), doc_id, fid))
    for nombre, original in guardar_archivos('archivos', EXT_DOC):
        db.execute('INSERT INTO archivos_documentos (documento_id,nombre_archivo,nombre_original) VALUES(?,?,?)',
                   (doc_id, nombre, original))
    db.commit()
    auditoria('actualizar', seccion, doc_id, 'Registro actualizado')
    flash('Registro actualizado.','success')
    return redirect(url_for('ficha_detalle', fid=fid))

@app.route('/fichas/<int:fid>/doc/<seccion>/eliminar', methods=['POST'])
@login_required
def doc_eliminar(fid, seccion):
    doc_id = request.form.get('doc_id','').strip()
    if doc_id:
        db = get_db()
        for a in db.execute('SELECT nombre_archivo FROM archivos_documentos WHERE documento_id=?',(doc_id,)):
            borrar_archivo(a['nombre_archivo'])
        db.execute('DELETE FROM documentos_hv WHERE id=? AND ficha_id=?',(doc_id,fid))
        db.commit()
        flash('Registro eliminado.','success')
    return redirect(url_for('ficha_detalle', fid=fid))

@app.route('/fichas/<int:fid>/doc/<seccion>/<int:doc_id>/archivos/<int:arch_id>/eliminar', methods=['POST'])
@login_required
def doc_archivo_eliminar(fid, seccion, doc_id, arch_id):
    db = get_db()
    a = db.execute('SELECT * FROM archivos_documentos WHERE id=? AND documento_id=?',(arch_id,doc_id)).fetchone()
    if a:
        borrar_archivo(a['nombre_archivo'])
        db.execute('DELETE FROM archivos_documentos WHERE id=?',(arch_id,))
        db.commit(); flash('Archivo eliminado.','success')
    return redirect(url_for('ficha_detalle', fid=fid))

# ─────────────────────────────────────────────
#  CATÁLOGOS (JSON API)
# ─────────────────────────────────────────────
CATS_VALIDAS = {'contrato','estamento','jornada','estado_tramite','tipo_documento'}

@app.route('/catalogo/<cat>/listar')
@login_required
def catalogo_listar(cat):
    if cat not in CATS_VALIDAS: return jsonify({'ok':False}),400
    db = get_db()
    items = db.execute('SELECT id,nombre FROM catalogos WHERE categoria=? ORDER BY nombre',(cat,)).fetchall()
    return jsonify([{'id':i['id'],'nombre':i['nombre']} for i in items])

@app.route('/catalogo/<cat>/agregar', methods=['POST'])
@login_required
def catalogo_agregar(cat):
    if cat not in CATS_VALIDAS: return jsonify({'ok':False}),400
    nombre = request.form.get('nombre','').strip()
    if not nombre: return jsonify({'ok':False,'error':'Escribe un nombre.'}),400
    db = get_db()
    ex = db.execute('SELECT id,nombre FROM catalogos WHERE categoria=? AND nombre=? COLLATE NOCASE',(cat,nombre)).fetchone()
    if ex: return jsonify({'ok':True,'id':ex['id'],'nombre':ex['nombre']})
    cur = db.execute('INSERT INTO catalogos (categoria,nombre) VALUES(?,?)',(cat,nombre))
    db.commit()
    return jsonify({'ok':True,'id':cur.lastrowid,'nombre':nombre})

@app.route('/catalogo/<cat>/<int:item_id>/eliminar', methods=['POST'])
@login_required
def catalogo_eliminar(cat, item_id):
    if cat not in CATS_VALIDAS: return jsonify({'ok':False}),400
    db = get_db()
    item = db.execute('SELECT nombre FROM catalogos WHERE id=? AND categoria=?',(item_id,cat)).fetchone()
    if not item: return jsonify({'ok':False,'error':'No encontrado.'}),404
    db.execute('DELETE FROM catalogos WHERE id=?',(item_id,)); db.commit()
    return jsonify({'ok':True})

# ─────────────────────────────────────────────
#  IMPORTAR EXCEL
# ─────────────────────────────────────────────
# Alias de encabezados para importar fichas — acepta cualquier variante
ALIAS_FICHAS = {
    'run (rut)':'rut','run':'rut','rut':'rut',
    'rol':'rol','pasaporte':'pasaporte',
    'país nacionalidad':'pais_nacionalidad','pais nacionalidad':'pais_nacionalidad','pais':'pais_nacionalidad',
    'nombre completo':'nombre','nombre':'nombre',
    'correo':'correo','correo electrónico':'correo','email':'correo',
    'región':'region','region':'region',
    'comuna':'comuna','ciudad':'ciudad',
    'fecha nacimiento':'fecha_nacimiento','fechanac':'fecha_nacimiento','fecha nac':'fecha_nacimiento',
    'estado civil':'estado_civil','estadocivil':'estado_civil',
    'jerarquía':'jerarquia','jerarquia':'jerarquia',
    'grado académico':'grado_academico','grado academico':'grado_academico','grado':'grado_academico',
    'título':'titulo','titulo':'titulo',
    'teléfono':'telefono','telefono':'telefono','tel':'telefono','fono':'telefono',
    'dirección':'direccion','direccion':'direccion','dir':'direccion',
    'observaciones':'observaciones','obs':'observaciones',
}

def _leer_excel(stream):
    from openpyxl import load_workbook
    wb = load_workbook(stream, data_only=True); ws = wb.active

    def s(v):
        if v is None: return ''
        if hasattr(v, 'strftime'): return v.strftime('%Y-%m-%d')
        return str(v).strip()

    # leer encabezados y mapear a campos internos usando alias
    primera = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    enc = [ALIAS_FICHAS.get(str(c).strip().lower(), str(c).strip().lower()) if c else '' for c in primera]

    personas = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in fila): continue
        persona = {}
        for i, clave in enumerate(enc):
            if clave and i < len(fila):
                persona[clave] = s(fila[i])
        if persona.get('nombre') or persona.get('rut'):
            personas.append(persona)
    return personas

@app.route('/fichas/importar', methods=['GET'])
@login_required
def fichas_importar_get():
    return render_template('importar.html', personas=None, error=None)

@app.route('/fichas/importar', methods=['POST'])
@login_required
def fichas_importar_post():
    archivo = request.files.get('excel')
    if not archivo or not archivo.filename:
        return render_template('importar.html', personas=None, error='Selecciona un archivo.')
    if not archivo.filename.lower().endswith(('.xlsx','.xls')):
        return render_template('importar.html', personas=None, error='Solo .xlsx o .xls.')
    try:
        personas = _leer_excel(archivo.stream)
        if not personas:
            return render_template('importar.html', personas=None, error='Sin datos o formato incorrecto.')
        return render_template('importar.html', personas=personas, error=None)
    except Exception as e:
        return render_template('importar.html', personas=None, error=f'Error leyendo el archivo: {e}')

@app.route('/fichas/confirmar-importacion', methods=['POST'])
@login_required
def fichas_confirmar_importacion():
    g = lambda k: request.form.get(k,'').strip()
    cantidad = int(g('cantidad_personas') or 0)
    db = get_db(); creadas=0; omitidas=[]
    campos = ['nombre','rut','rol','pasaporte','pais_nacionalidad','region','comuna','ciudad',
              'correo','fecha_nacimiento','estado_civil','jerarquia','grado_academico',
              'titulo','telefono','direccion','observaciones']
    for i in range(cantidad):
        nombre = g(f'nombre_{i}')
        rut    = g(f'rut_{i}')
        if not nombre: continue
        if not rut:
            omitidas.append(f'{nombre} (RUT vacío)'); continue
        if db.execute('SELECT id FROM fichas WHERE rut=?',(rut,)).fetchone():
            omitidas.append(f'{nombre} (RUT ya existe)'); continue
        numeros_usados = {r[0] for r in db.execute('SELECT numero_ficha FROM fichas WHERE numero_ficha IS NOT NULL')}
        num = 1
        while num in numeros_usados: num += 1
        vals = {c: g(f'{c}_{i}') for c in campos}
        db.execute('''INSERT INTO fichas
            (numero_ficha,nombre,rut,rol,pasaporte,pais_nacionalidad,region,comuna,ciudad,
             correo,fecha_nacimiento,estado_civil,jerarquia,grado_academico,titulo,
             telefono,direccion,observaciones,activo)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)''',
            (num,vals['nombre'],vals['rut'],vals['rol'],vals['pasaporte'],vals['pais_nacionalidad'],
             vals['region'],vals['comuna'],vals['ciudad'],vals['correo'],vals['fecha_nacimiento'],
             vals['estado_civil'],vals['jerarquia'],vals['grado_academico'],vals['titulo'],
             vals['telefono'],vals['direccion'],vals['observaciones']))
        db.commit(); creadas+=1
        auditoria('crear','ficha',db.execute('SELECT last_insert_rowid()').fetchone()[0],
                  f'Importada desde Excel: {nombre} ({rut})')
    msg = f'{creadas} hoja(s) de vida importada(s) correctamente.'
    if omitidas: msg += f' Omitidas: {"; ".join(omitidas)}'
    flash(msg, 'error' if omitidas else 'success')
    return redirect(url_for('fichas_listado'))




# ─────────────────────────────────────────────
#  IMPORTAR SECCIONES DESDE EXCEL
# ─────────────────────────────────────────────
# Mapeo nombre de hoja → seccion interna
NOMBRE_A_SECCION = {
    'designaciones':                  'designaciones',
    'cese de funciones':              'cese_funciones',
    'responsabilidad administrativa': 'responsabilidad_administrativa',
    'inhabilidades':                  'inhabilidades',
    'estudios':                       'estudios',
    'contrato a honorarios':          'contrato_honorarios',
    'licencias medicas':              'licencias_medicas',
    'licencias médicas':              'licencias_medicas',
    'permisos y feriados':            'permisos_feriados',
    'comisiones y becas':             'comisiones_becas',
    'cargas familiares':              'cargas_familiares',
    'calificacion':                   'calificacion',
    'calificación':                   'calificacion',
    'destinaciones':                  'destinaciones',
    'ceses':                          'ceses',
    'anotaciones':                    'anotaciones',
    'declaraciones de patrimonio':    'declaraciones_patrimonio',
    'modifica rectifica':             'modifica_rectifica',
}

# Campos que van en la tabla principal (el resto va a datos_extra JSON)
CAMPOS_BASE_HV = {'tipo_documento', 'numero_documento', 'fecha_documento',
                  'servicio_documento', 'estado_tramite'}

# Mapeo flexible de encabezados comunes a nombres internos
ALIAS_ENCABEZADOS = {
    'run':                  'rut',
    'rut':                  'rut',
    'tipo documento':       'tipo_documento',
    'tipo doc':             'tipo_documento',
    'n° documento':         'numero_documento',
    'numero documento':     'numero_documento',
    'nro documento':        'numero_documento',
    'n documento':          'numero_documento',
    'fecha documento':      'fecha_documento',
    'servicio':             'servicio_documento',
    'servicio documento':   'servicio_documento',
    'servicio desempeño':   'servicio_documento',
    'estado tramite':       'estado_tramite',
    'estado trámite':       'estado_tramite',
}

def _normalizar_enc(nombre):
    """Convierte encabezado a clave interna si tiene alias, si no lo normaliza."""
    n = nombre.lower().strip()
    if n in ALIAS_ENCABEZADOS:
        return ALIAS_ENCABEZADOS[n]
    # normalización básica: minúsculas, sin tildes, guión bajo por espacio
    for a, b in [('á','a'),('é','e'),('í','i'),('ó','o'),('ú','u'),('ñ','n'),(' ','_'),('/','_'),('-','_')]:
        n = n.replace(a, b)
    return n.strip('_')

def _leer_secciones_excel(stream):
    from openpyxl import load_workbook
    wb = load_workbook(stream, data_only=True)

    def s(v):
        if v is None: return ''
        if hasattr(v, 'strftime'): return v.strftime('%Y-%m-%d')
        return str(v).strip()

    resultado = {}
    for nombre_hoja in wb.sheetnames:
        # buscar qué sección corresponde a esta hoja
        seccion_id = None
        nombre_lower = nombre_hoja.lower().strip()
        for k, v in NOMBRE_A_SECCION.items():
            if k in nombre_lower or nombre_lower in k:
                seccion_id = v; break
        if not seccion_id:
            continue

        ws = wb[nombre_hoja]
        filas_iter = ws.iter_rows(min_row=1, values_only=True)

        # primera fila = encabezados
        primera = next(filas_iter, None)
        if not primera: continue
        enc = [_normalizar_enc(str(c).strip()) if c else '' for c in primera]

        filas = []
        for fila in filas_iter:
            if not any(c is not None for c in fila): continue

            # mapear cada celda a su clave normalizada
            registro = {}
            for i, clave in enumerate(enc):
                if not clave: continue
                registro[clave] = s(fila[i] if i < len(fila) else None)

            # el RUT puede estar como 'rut' o 'run'
            rut = registro.get('rut') or registro.get('run') or ''
            if not rut: continue
            registro['rut'] = rut

            filas.append({'seccion': seccion_id, 'rut': rut, 'datos': registro})

        if filas:
            resultado[nombre_hoja] = {'seccion_id': seccion_id, 'filas': filas}

    return resultado


@app.route('/fichas/importar-secciones', methods=['GET'])
@login_required
def importar_secciones_get():
    return render_template('importar_secciones.html', resultado=None, error=None)


@app.route('/fichas/importar-secciones', methods=['POST'])
@login_required
def importar_secciones_post():
    archivo = request.files.get('excel')
    if not archivo or not archivo.filename:
        return render_template('importar_secciones.html', resultado=None, error='Selecciona un archivo.')
    if not archivo.filename.lower().endswith(('.xlsx', '.xls')):
        return render_template('importar_secciones.html', resultado=None, error='Solo .xlsx o .xls.')
    try:
        resultado = _leer_secciones_excel(archivo.stream)
        if not resultado:
            return render_template('importar_secciones.html', resultado=None,
                error='No se encontraron hojas reconocidas. Verifica que los nombres coincidan (ej: Designaciones, Estudios...)')
        return render_template('importar_secciones.html', resultado=resultado, error=None)
    except Exception as e:
        return render_template('importar_secciones.html', resultado=None, error=f'Error leyendo el archivo: {e}')


@app.route('/fichas/confirmar-importacion-secciones', methods=['POST'])
@login_required
def confirmar_importacion_secciones():
    db    = get_db()
    total = 0
    no_encontrados = []
    cantidad = int(request.form.get('cantidad_registros', 0))

    for i in range(cantidad):
        rut     = request.form.get(f'r_{i}_rut', '').strip()
        seccion = request.form.get(f'r_{i}_seccion', '').strip()
        if not rut or not seccion: continue

        ficha = db.execute('SELECT id FROM fichas WHERE rut=?', (rut,)).fetchone()
        if not ficha:
            if rut not in no_encontrados:
                no_encontrados.append(rut)
            continue

        # separar campos base de extras
        datos = json.loads(request.form.get(f'r_{i}_datos', '{}'))
        base  = {k: datos.pop(k, '') for k in list(CAMPOS_BASE)}
        # quitar rut y seccion de extra
        datos.pop('rut', None); datos.pop('seccion', None)
        datos.pop('run', None)

        db.execute("""INSERT INTO documentos_hv
            (ficha_id, seccion, tipo_documento, numero_documento, fecha_documento,
             servicio_documento, estado_tramite, datos_extra)
            VALUES (?,?,?,?,?,?,?,?)""",
            (ficha['id'], seccion,
             base.get('tipo_documento',''), base.get('numero_documento',''),
             base.get('fecha_documento',''), base.get('servicio_documento',''),
             base.get('estado_tramite',''), json.dumps(datos)))
        total += 1

    db.commit()
    msg = f'{total} registro(s) importado(s) correctamente.'
    if no_encontrados:
        msg += f' RUTs no encontrados: {", ".join(no_encontrados)}'
    flash(msg, 'error' if no_encontrados else 'success')
    return redirect(url_for('fichas_listado'))


# ─────────────────────────────────────────────
#  ARRANQUE
# ─────────────────────────────────────────────
init_db()

if __name__ == '__main__':
    app.run(debug=True)
